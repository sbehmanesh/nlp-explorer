from django.shortcuts import render
from django.http import HttpResponse
import spacy
import openpyxl
from .models import Festival,Features
from django.db.models import Q
from django.shortcuts import render
from django.db.models import Count
from django.views.decorators.csrf import csrf_protect 
import re

list_of_monthes = [
    'january',
    'march',
    'may',
    'july',
    'september',
    'november',
    'february',
    'april',
    'june',
    'august',
    'october',
    'december' 
]

list_of_short_monthes = [
    'jan',
    'mar',
    'may',
    'jul',
    'sep',
    'nov',
    'feb',
    'apr',
    'jun',
    'aug',
    'oct',
    'dec' 
]

blacklist_words = [
    'and',
    'or'
]

def index(request):
    return render(request, 'search/index.html')
    

@csrf_protect
def search(request):
    view_data = {}
    if request.method == "POST":

        main_string = request.POST.get('string')

        clean_string = clean_string_from_non_alpha_numeric_chars(main_string)

        strings = split_string_according_to_prepositions(clean_string)

        complete_query = Q()
        keywords = ""
        list_of_types = []
        

        for and_string in strings:

            inter_statements_query = Q()
            or_strings = and_string.split("and")

            for string in or_strings:
                locs = []
                dates = []
                titles = []
                types = []
                string = string.strip()
                caches = get_features(string)
                for cache in caches:
                    if cache.word not in blacklist_words:
                        word = clean_string_from_non_alpha_numeric_chars(cache.word)
                        if cache.type == 'title':
                            if word not in titles:
                                titles.append(word)
                        if cache.type == 'address':
                            if word not in locs:
                                locs.append(word)
                        if cache.type == 'date':
                            if word not in dates:
                                dates.append(word)
                        if cache.type == 'type':
                            if word not in types:
                                types.append(word)

                nlp_locs,nlp_dates,nlp_titles = extract_locations_and_dates(string)
                all_features_cached = locs + dates + titles
                for word in nlp_locs:
                    if word not in all_features_cached:
                        locs.append(word) 
                for word in nlp_dates:
                    if word not in all_features_cached:
                        dates.append(word) 
                for word in nlp_titles:
                    if word not in all_features_cached:
                        titles.append(word) 

                date_prim = month_search(string)
                for month in date_prim:
                    if month not in dates:
                        dates.append(month)
                

                loc_query = Q()
                for loc in locs:
                    loc_query = loc_query & Q(address__icontains=loc)
                if len(locs) != 0:
                    keywords = keywords + " , " + ' , '.join(locs)

                date_query = Q()
                for date in dates:
                    date_query = date_query & Q(date__icontains=date)
                if len(dates) != 0:
                    keywords = keywords + " , " + ' , '.join(dates)

                title_query = Q()
                for title in titles:
                    title_query = title_query & (Q(title__icontains=title) | Q(type__icontains=title))
                if len(titles) != 0:
                    keywords = keywords + " , " + ' , '.join(titles)
                
                type_query = Q()
                for type in types:
                    type_query = type_query & (Q(type__icontains=type) | Q(title__icontains=type))
                if len(types) != 0:
                    keywords = keywords + " , " + ' , '.join(types)

                final_query = loc_query & date_query & title_query & type_query
                inter_statements_query = inter_statements_query | final_query

            complete_query = complete_query & inter_statements_query
        
        
        find_all_types = Festival.objects.filter(complete_query).annotate(total_count=Count('id'))
                
        for special_type in find_all_types:
            if special_type.type not in list_of_types:
                list_of_types.append(special_type.type)

        # print(Festival.objects.filter(complete_query).query)
        view_data['all_results'] = Festival.objects.filter(complete_query)
        view_data['string'] = main_string
        view_data['keywords'] = keywords[3:]
        view_data['types'] = list_of_types

    return render(request, 'search/result.html', view_data)


def split_string_according_to_prepositions(string):
    exploded_by_in = string.split("in")
    and_strings = []
    for sub_string in exploded_by_in:
        exploded_by_at = sub_string.strip().split("at")
        and_strings = and_strings + exploded_by_at

    return and_strings


def clean_string_from_non_alpha_numeric_chars(string):
    # clean input string from non alpha-numeric characters
    pattern = r'(?<!\d)\d{2}(?!\d)'
    no_2digit_number_in_string = re.sub(pattern, '', string)
    clean_string = filter(lambda x: x.isalnum() or x.isspace(), no_2digit_number_in_string)
    clean_string = "".join(clean_string)
    return clean_string


def extract_locations_and_dates(string):
    nlp_wk = spacy.load('en_core_web_sm')
    doc = nlp_wk(string)
    locations = []
    dates = []
    title = []
    for chunk in doc.ents:
        if chunk.label_ in ['LOC','GPE']:
            splited = chunk.text.split(' ')
            locations.extend(splited)
        elif chunk.label_ == 'DATE':
            splited = chunk.text.split(' ')
            dates.extend(splited)
        else:
            splited = chunk.text.split(' ')
            title.extend(splited)

    return locations,dates,title


def get_features(string):
    words = string.split()
    cached_features = []
    for word in words:
        if len(word) > 2 :
            cache = Features.objects.filter(word__istartswith=word).first()
            if cache :
                cached_features.append(cache)

    return cached_features


def upload_excel_to_db():
    Festival.objects.all().delete()
    
    excel_path = "static/data.xlsx"
    wb_obj = openpyxl.load_workbook(excel_path)
    sheet_obj = wb_obj.active
    row_num = sheet_obj.max_row

    for i in range(2,row_num):
        if sheet_obj.cell(row = i, column = 1).value:
            Festival(
                title = sheet_obj.cell(row = i, column = 1).value,
                date = sheet_obj.cell(row = i, column = 2).value,
                address = sheet_obj.cell(row = i, column = 3).value,
                lat = sheet_obj.cell(row = i, column = 4).value,
                long = sheet_obj.cell(row = i, column = 5).value,
                link = sheet_obj.cell(row = i, column = 6).value,
                type = sheet_obj.cell(row = i, column = 7).value
                ).save()


def extract_features_from_data():
    all_data = Festival.objects.all()
    already_added_title = []
    already_added_address = []
    already_added_date = []
    already_added_type = []

    for data in all_data:
        words = (data.title).split()
        for word in words:
            if len(word) > 2:
                if word not in already_added_title:
                    already_added_title.append(word)
                    Features(word=word,type='title').save()

        words = (data.address).split()
        for word in words:
            if len(word) > 2:
                if word not in already_added_address:
                    already_added_address.append(word)
                    Features(word=word,type='address').save()

        words = (data.date).split()
        for word in words:
            if len(word) > 2:
                if word not in already_added_date:
                    already_added_date.append(word)
                    Features(word=word,type='date').save()
        
        words = (data.type).split()
        for word in words:
            if len(word) > 2:
                if word not in already_added_type:
                    already_added_type.append(word)
                    Features(word=word,type='type').save()

    # print('number of title features added : '  ,len(already_added_title))
    # print('number of address features added : ',len(already_added_address))
    # print('number of date features added : '   ,len(already_added_date))
    # print('number of type features added : '   ,len(already_added_type))


def month_search(string):
    words = string.split()
    monthes = []
    for word in words:
        if word.lower() in list_of_monthes:
            index = list_of_monthes.index(word.lower())
            monthes.append(list_of_short_monthes[index])
        elif word.lower() in list_of_short_monthes:
            index = list_of_short_monthes.index(word.lower())
            monthes.append(list_of_monthes[index])
    
    return monthes